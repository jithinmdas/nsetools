"""
Author: Jithin M
Email: jmd@jithinmdas.com
This program is used for equity analysis. It is mainly for indices.
"""

import sys
import time
import os.path
from openpyxl import load_workbook
from os import write
from numpy import equal
from pandas.core.frame import DataFrame
from requests import NullHandler
import xlsxwriter
import pandas as pd
from nsepy import get_history
from nsepy import get_index_pe_history
from datetime import date, datetime, timedelta


class NseData:
    def __init__(self, symbol, start_date, end_date) -> None:
        self.symbol = symbol
        self.start_date = start_date
        self.end_date = end_date

    def get_price(self):
        try:
            data = get_history(self.symbol, self.start_date,
                               self.end_date, index=True)
        except Exception:
            sys.exit('Failed to retrieve price')
        print(data)
        return data

    def get_pe_pb_div(self):
        try:
            data = get_index_pe_history(
                self.symbol, self.start_date, self.end_date)
        except Exception:
            sys.exit('Failed to retrieve pe_pb_div')
        print(data)
        return data


def store_data_xlsx(filename, index, result):
    writer = pd.ExcelWriter(filename+index+'.xlsx', engine='xlsxwriter',
                            date_format='dd mmm yyyy')
    workbook = writer.book
    num_format = workbook.add_format({'num_format': '#,##0.00'})
    perc_format = workbook.add_format({'num_format': '##0.00%'})
    date_format = workbook.add_format({'num_format': 'dd mmm yyyy'})

    result.to_excel(writer, sheet_name=index)
    worksheet = writer.sheets[index]
    worksheet.set_column('A:A', 18, date_format)
    worksheet.set_column('B:O', None, num_format)
    worksheet.set_column('P:P', None, perc_format)
    worksheet.freeze_panes(1, 0)
    # chart = workbook.add_chart({'type': 'line'})
    # chart.add_series({'values': '=$P$2:$P$1048576'})
    # worksheet.insert_chart('D2', chart)
    worksheet.conditional_format('P1:P1048576', {'type': '3_color_scale',
                                                 'min_type': 'percent',
                                                 'mid_type': 'percent',
                                                 'max_type': 'percent'})
    writer.save()


def update_mean_till_date(dataframe, column_to_refer, column_to_add):
    for itr in range(1, len(dataframe)+1):
        sub_mean = dataframe[[column_to_refer]].head(itr).mean(skipna=True)
        dataframe[column_to_add].iat[itr-1] = sub_mean


def conditional_formatting(dataframe, column_to_refer):
    dataframe.style


def read_excel(filename, sheetname):
    wb = load_workbook(filename+sheetname+'.xlsx', read_only=True)
    if sheetname in wb.sheetnames:
        try:
            read_data = pd.read_excel(filename+sheetname+'.xlsx', sheet_name=sheetname,
                                      header=0, index_col=0)
        except Exception as err:
            print(err)
            return None
        print('Data read')
        read_data_end_date = read_data.index[-1]
        start_date = datetime.date(read_data_end_date + timedelta(days=1))
        return [read_data, start_date]
    else:
        return [None, None]


def get_data(nse):
    price = nse.get_price()
    if price is None:
        sys.exit('Failed to retrieve price')

    time.sleep(10)

    pe_pb_div = nse.get_pe_pb_div()
    if pe_pb_div is None:
        sys.exit('Failed to retrieve pe_pb_div')

    # Concat two tables
    result = pd.concat([price, pe_pb_div], axis=1, join="inner")
    return result


def create_data_by_index(filename, start_date, end_date, index_name):
    # Check if file already exists
    if os.path.isfile(filename + index_name + '.xlsx'):
        print("File exists")
        [read_data, tmp_start_date] = read_excel(filename, index_name)
        # read_data_start_date = read_data.index[0]
        # print(read_data_start_date)
        if tmp_start_date == None:
            print("Data not read for " + index_name)
        print(start_date)
        print(end_date)
    else:
        print('Creating new file')

    nse = NseData(index_name, start_date, end_date)

    # Get the tables
    result = get_data(nse)
    print(result)

    if os.path.isfile(filename):
        # Append to current table
        result = pd.concat([read_data, result], axis=0)

    # Add new columns
    result['Average P/E'] = result['P/E']
    result['Average P/B'] = result['P/B']
    result['Average Div'] = result['Div Yield']
    result['EPS'] = result['Close']/result['P/E']

    # Find mean till date and update the columns
    update_mean_till_date(result, 'P/E', 'Average P/E')
    update_mean_till_date(result, 'P/B', 'Average P/B')
    update_mean_till_date(result, 'Div Yield', 'Average Div')

    result['Fair price'] = result['Average P/E']*result['EPS']
    result['Diff in percent'] = \
        (result['Close'] - result['Fair price'])/(result['Fair price'])

    # conditional_formatting(result, 'Diff in percent')
    print(result)
    return result


def main():
    filename = 'nsedata'
    start_date = date(1999, 1, 1)
    end_date = date.today() - timedelta(days=1)
    indices = {'NIFTY 50': '',
               'NIFTY Next 50': '',
               'NIFTY Auto': '',
               'NIFTY Bank': '',
               'NIFTY FMCG': ''}

    for index in sorted(indices):
        indices[index] = create_data_by_index(
            filename, start_date, end_date, index)
        store_data_xlsx(filename, index, indices[index])


if __name__ == "__main__":
    main()
